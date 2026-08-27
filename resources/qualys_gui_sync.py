#!/usr/bin/env python3
"""Sync Qualys business units and users from an xlsx workbook via the GUI session API.

This is the companion script for the two objects the Terraform provider cannot
manage (no API exists for them): business units and user accounts. It drives the
same legacy /fo/options/*.php endpoints the Qualys web UI uses, which require a
browser-style session (a login cookie plus a per-form CSRF token), so it cannot
ride the provider's stateless Basic-auth client.

It reads two named Excel tables from the workbook:

  qualys_business_unit
    enabled, key, title, asset_groups__list, comments, delete
  qualys_user
    enabled, key, first_name, last_name, title, phone, fax, email,
    address_1, address_2, city, country, state, zip_code, language,
    date_format, time_zone, user_role, business_unit_title, gui, api,
    manage_vm, modify_remedy_policy, add_vhost, add_asset, option_profile,
    purge_host, modify_ntauth, manage_compliance, approve_exceptions,
    modify_compliance_policy, create_user_defined_control,
    modify_all_user_defined_control, manage_webapp, create_webapps,
    latest_vulnerabilities, scan_complete_notification, scan_notification,
    map_notification, report_notification, exception_notification,
    user_session_timeout, delete

Idempotency is tracked in a state file (default ./qualys_gui_state.json) that
maps business-unit title -> id and user email -> id. On the first run objects
are created and their ids recorded; on later runs the recorded id is used to
update in place. Objects whose row has `delete` set to a truthy value are
deleted and removed from state. Because the state file is the source of truth
for ids, objects created or deleted out-of-band in the Qualys UI are not
detected; see the README note on this limitation.

The new object id is extracted from the create response by scanning for
`edit=<id>` in the returned HTML. This is a best-effort heuristic against the
GUI's response format; if a create succeeds but no id is found the script
raises an error rather than guessing. Adjust the regex in _save_business_unit /
_save_user if your platform's response differs.

Business-unit `asset_groups` is a list of *titles* that is resolved to ids
before submission via the VMDR API (Basic auth) when QUALYS_API_URL is set;
otherwise the titles are passed through as-is.

The user-to-business-unit relationship is established by the user's
`business_unit_title` field (a user belongs to a business unit), so business
units are synced first and users reference them by title.

Environment:
  QUALYS_GUI_URL     GUI host, e.g. https://qualysguard.qg2.apps.qualys.eu
  QUALYS_USERNAME    GUI login username
  QUALYS_PASSWORD    GUI login password
  QUALYS_API_URL     (optional) API host for asset-group title resolution,
                     e.g. https://qualysapi.qualys.eu

Usage:
  python qualys_gui_sync.py <workbook.xlsx> [--state state.json] [--dry-run]
"""

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

GUI_URL = os.environ.get("QUALYS_GUI_URL", "").rstrip("/")
USERNAME = os.environ.get("QUALYS_USERNAME", "")
PASSWORD = os.environ.get("QUALYS_PASSWORD", "")
API_URL = os.environ.get("QUALYS_API_URL", "").rstrip("/")

LOGIN_PATH = "/fo/user_login.php"
BU_EDIT_PATH = "/fo/options/bizunit_edit.php"
BU_DELETE_PATH = "/fo/options/delete_bizunit.php"
USER_EDIT_PATH = "/fo/options/user_edit.php"
USER_DELETE_PATH = "/fo/options/delete_user.php"

# The GUI form pages embed a session-bound CSRF token in a hidden field.
_FORM_ID_RE = re.compile(r'name=["\']_form_id["\'][^>]*value=["\']([^"\']+)["\']', re.I)
_FORM_ID_RE_ALT = re.compile(r'value=["\']([^"\']+)["\'][^>]*name=["\']_form_id["\']', re.I)

# VMDR asset-group list response: <ASSET_GROUP><ID>..</ID><TITLE>..</TITLE>...
_ASSET_GROUP_RE = re.compile(
    r"<ASSET_GROUP>\s*<ID>(\d+)</ID>\s*<TITLE>(.*?)</TITLE>", re.S
)


class SyncError(Exception):
    pass


# ---------------------------------------------------------------------------
# GUI session
# ---------------------------------------------------------------------------


class QualysGUISession:
    """A browser-style session against the Qualys GUI host."""

    def __init__(self, base_url, username, password):
        self.base_url = base_url
        self.username = username
        self.password = password
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": "terraform-provider-qualys-gui-sync",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            }
        )

    def login(self):
        if not (self.base_url and self.username and self.password):
            raise SyncError(
                "QUALYS_GUI_URL, QUALYS_USERNAME and QUALYS_PASSWORD must be set"
            )
        url = self.base_url + LOGIN_PATH
        # Match the browser's login request: an XHR POST with form-urlencoded
        # body. The x-nonce returned in the response headers is a CSP nonce, not
        # a token the login must submit, so it is not sent.
        self.session.headers.update(
            {
                "X-Requested-With": "XMLHttpRequest",
                "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
            }
        )
        data = {
            "UserLogin": self.username,
            "UserPasswd": self.password,
            "_form_action1": "Please wait...",
            "_form_action": "Login",
            "_form_visited": "1",
            "timezone": "UTC",
            "gmtoffset": "0",
            "timezone_abbr": "UTC",
            "is_dst": "0",
        }
        resp = self.session.post(
            url,
            data=data,
            allow_redirects=True,
            timeout=60,
        )
        if "QualysSession" not in self.session.cookies:
            raise SyncError(
                f"login to {url} did not establish a QualysSession cookie "
                f"(HTTP {resp.status_code}); the GUI login form may differ on this "
                "platform - adjust LOGIN_PATH/fields in qualys_gui_sync.py"
            )
        return resp

    def get_form_id(self, url):
        """GET a form page and return its session-bound _form_id CSRF token."""
        resp = self.session.get(url, timeout=60)
        for pattern in (_FORM_ID_RE, _FORM_ID_RE_ALT):
            m = pattern.search(resp.text)
            if m:
                return m.group(1)
        raise SyncError(f"could not find _form_id on {url} (HTTP {resp.status_code})")

    def post_form(self, url, data):
        """POST form data, returning the response text."""
        resp = self.session.post(url, data=data, allow_redirects=True, timeout=120)
        if resp.status_code >= 400:
            raise SyncError(f"POST {url} returned HTTP {resp.status_code}")
        return resp.text


# ---------------------------------------------------------------------------
# VMDR asset-group resolution (optional)
# ---------------------------------------------------------------------------


def resolve_asset_groups(titles):
    """Resolve asset-group titles to ids via the VMDR API (Basic auth).

    Returns a dict title -> id. If QUALYS_API_URL is not set, returns an empty
    dict (callers then pass titles through unresolved).
    """
    if not API_URL:
        return {}
    if not (USERNAME and PASSWORD):
        raise SyncError("QUALYS_API_URL set but QUALYS_USERNAME/PASSWORD missing")
    url = API_URL + "/api/2.0/fo/asset/group/"
    resp = requests.post(
        url,
        data={"action": "list"},
        auth=(USERNAME, PASSWORD),
        headers={"X-Requested-With": "terraform-provider-qualys-gui-sync"},
        timeout=60,
    )
    if resp.status_code >= 400:
        raise SyncError(f"asset group list returned HTTP {resp.status_code}")
    found = {}
    for gid, gtitle in _ASSET_GROUP_RE.findall(resp.text):
        found[gtitle.strip()] = gid
    return found


# ---------------------------------------------------------------------------
# xlsx reading
# ---------------------------------------------------------------------------


def read_table(wb, table_name):
    """Return a list of dicts for a named Excel table, honouring `enabled`."""
    for ws in wb.worksheets:
        if table_name not in ws.tables:
            continue
        ref = ws.tables[table_name].ref
        rows = ws[ref]
        headers = [str(c.value).strip() for c in rows[0]]
        out = []
        for row in rows[1:]:
            raw = dict(zip(headers, [c.value for c in row]))
            enabled = raw.get("enabled")
            if enabled not in (True, 1, "TRUE", "True", "true", "YES", "Yes", "yes"):
                continue
            out.append(raw)
        return out
    return []


def split_list(value):
    if value is None:
        return []
    if isinstance(value, str):
        return [v.strip() for v in value.split(";") if v.strip()]
    return [value]


def truthy(value):
    return value in (True, 1, "TRUE", "True", "true", "YES", "Yes", "yes", "1")


# ---------------------------------------------------------------------------
# Business units
# ---------------------------------------------------------------------------


def bu_payload(row, asset_group_ids, form_id, edit_id):
    return {
        "business_unit_title": row.get("title") or "",
        "asset_groups": "|".join(asset_group_ids),
        "users": "",
        "comments": row.get("comments") or "",
        "edit": edit_id,
        "tab_selected": "",
        "subtab_selected": "",
        "save_type_on_confirm": "0",
        "wf": "",
        "in_popup": "",
        "refresh_parent": "bu_page",
        "_show_changes": "",
        "ext_edit": "1",
        "helpUrl": "https://docs.qualys.com/en/vm/business_units/win_business_unit.htm",
        "_form_action": "Save",
        "_form_visited": "1",
        "_page_title": "Edit Business Unit" if edit_id else "New Business Unit",
        "_form_id": form_id,
    }


def sync_business_units(session, wb, state, dry_run):
    rows = read_table(wb, "qualys_business_unit")
    if not rows:
        print("qualys_business_unit: no enabled rows")
        return

    asset_group_map = resolve_asset_groups(
        {t for r in rows for t in split_list(r.get("asset_groups__list"))}
    )

    for row in rows:
        title = (row.get("title") or "").strip()
        if not title:
            print("  skip: business unit row with no title")
            continue
        key = title
        existing_id = state.get("business_units", {}).get(key)

        if truthy(row.get("delete")):
            if existing_id:
                if dry_run:
                    print(f"  [dry-run] delete business unit '{title}' ({existing_id})")
                else:
                    _delete_business_unit(session, existing_id)
                    state["business_units"].pop(key, None)
                    print(f"  deleted business unit '{title}' ({existing_id})")
            else:
                print(f"  skip delete '{title}': not in state")
            continue

        asset_group_ids = []
        for t in split_list(row.get("asset_groups__list")):
            gid = asset_group_map.get(t)
            asset_group_ids.append(gid if gid else t)

        if existing_id:
            if dry_run:
                print(f"  [dry-run] update business unit '{title}' ({existing_id})")
            else:
                _save_business_unit(session, row, asset_group_ids, existing_id)
                print(f"  updated business unit '{title}' ({existing_id})")
        else:
            if dry_run:
                print(f"  [dry-run] create business unit '{title}'")
            else:
                new_id = _save_business_unit(session, row, asset_group_ids, "0")
                state.setdefault("business_units", {})[key] = new_id
                print(f"  created business unit '{title}' ({new_id})")


def _save_business_unit(session, row, asset_group_ids, edit_id):
    url = session.base_url + BU_EDIT_PATH
    form_id = session.get_form_id(url + f"?edit={edit_id}&refresh_parent=bu_page&ext_edit=1")
    payload = bu_payload(row, asset_group_ids, form_id, edit_id)
    text = session.post_form(url, payload)
    # The create response carries the new id in the redirect/refresh URL.
    m = re.search(r"edit=(\d+)", text)
    if edit_id == "0":
        if not m:
            raise SyncError(f"business unit create did not return an id: {text[:300]}")
        return m.group(1)
    return edit_id


def _delete_business_unit(session, bu_id):
    url = session.base_url + BU_DELETE_PATH
    form_id = session.get_form_id(url)
    session.post_form(
        url,
        {
            "checkbox[1]": bu_id,
            "confirmed": "yes",
            "_form_id": form_id,
            "dl[name]": "",
            "dl[date]": "",
            "dl[offset]": "0",
            "dl[no_update]": "0",
            "dl[hide_columns]": "",
            "dl[column]": "name",
            "dl[order]": "2",
            "dl[rows_shown]": "20",
            "dl[row_selected]": "",
            "dl[quick_filter]": "",
            "dl[export]": "",
            "dl[row_select_available]": bu_id,
            "search_hid": "0",
            "search_on": "",
            "dl_action": "",
        },
    )


# ---------------------------------------------------------------------------
# Users
# ---------------------------------------------------------------------------


def user_payload(row, business_unit_id, form_id, edit_id):
    return {
        "first_name": row.get("first_name") or "",
        "last_name": row.get("last_name") or "",
        "title": row.get("title") or "",
        "phone": row.get("phone") or "",
        "fax": row.get("fax") or "",
        "e_mail_address": row.get("email") or "",
        "address_1": row.get("address_1") or "",
        "address_2": row.get("address_2") or "",
        "city": row.get("city") or "",
        "country": row.get("country") or "",
        "state": row.get("state") or "",
        "zip_code": row.get("zip_code") or "",
        "external_id": "",
        "api_external_id": "",
        "language": row.get("language") or "en",
        "date_format": row.get("date_format") or "dmy",
        "time_zone": row.get("time_zone") or "",
        "user_role": row.get("user_role") or "",
        "business_unit": business_unit_id,
        "gui": "1" if truthy(row.get("gui", True)) else "0",
        "api": "1" if truthy(row.get("api", True)) else "0",
        "manage_vm": "1" if truthy(row.get("manage_vm", True)) else "0",
        "modify_remedy_policy": "1" if truthy(row.get("modify_remedy_policy")) else "0",
        "add_vhost": "1" if truthy(row.get("add_vhost")) else "0",
        "add_asset": "1" if truthy(row.get("add_asset")) else "0",
        "option_profile": "1" if truthy(row.get("option_profile")) else "0",
        "purge_host": "1" if truthy(row.get("purge_host")) else "0",
        "modify_ntauth": "1" if truthy(row.get("modify_ntauth")) else "0",
        "manage_compliance": "1" if truthy(row.get("manage_compliance")) else "0",
        "approve_exceptions": "1" if truthy(row.get("approve_exceptions")) else "0",
        "modify_compliance_policy": "1" if truthy(row.get("modify_compliance_policy")) else "0",
        "create_user_defined_control": "1" if truthy(row.get("create_user_defined_control")) else "0",
        "modify_all_user_defined_control": "1" if truthy(row.get("modify_all_user_defined_control")) else "0",
        "manage_webapp": "1" if truthy(row.get("manage_webapp")) else "0",
        "create_webapps": "1" if truthy(row.get("create_webapps")) else "0",
        "latest_vulnerabilities": "1" if truthy(row.get("latest_vulnerabilities")) else "0",
        "scan_complete_notification": "1" if truthy(row.get("scan_complete_notification")) else "0",
        "scan_notification": "1" if truthy(row.get("scan_notification")) else "0",
        "map_notification": "1" if truthy(row.get("map_notification")) else "0",
        "report_notification": "1" if truthy(row.get("report_notification")) else "0",
        "exception_notification": "1" if truthy(row.get("exception_notification")) else "0",
        "no_of_days": "01",
        "user_session_timeout": row.get("user_session_timeout") or "30",
        "edit": edit_id,
        "tab_selected": "",
        "subtab_selected": "",
        "parent_opened": "",
        "in_popup": "",
        "refresh_parent": "1",
        "_show_changes": "",
        "api_external_id_message": "",
        "include_disable": "",
        "info_warning_txt": "",
        "div_selected": "",
        "ext_edit": "1",
        "helpUrl": "https://docs.qualys.com/en/vm/user_accounts/win_user.htm",
        "_form_action": "Save",
        "_form_visited": "1",
        "_page_title": "Edit User" if edit_id else "New User",
        "_form_id": form_id,
    }


def sync_users(session, wb, state, dry_run):
    rows = read_table(wb, "qualys_user")
    if not rows:
        print("qualys_user: no enabled rows")
        return

    for row in rows:
        email = (row.get("email") or "").strip()
        if not email:
            print("  skip: user row with no email")
            continue
        key = email
        existing_id = state.get("users", {}).get(key)

        if truthy(row.get("delete")):
            if existing_id:
                if dry_run:
                    print(f"  [dry-run] delete user '{email}' ({existing_id})")
                else:
                    _delete_user(session, existing_id)
                    state["users"].pop(key, None)
                    print(f"  deleted user '{email}' ({existing_id})")
            else:
                print(f"  skip delete '{email}': not in state")
            continue

        bu_title = (row.get("business_unit_title") or "").strip()
        bu_id = state.get("business_units", {}).get(bu_title, bu_title)

        if existing_id:
            if dry_run:
                print(f"  [dry-run] update user '{email}' ({existing_id})")
            else:
                _save_user(session, row, bu_id, existing_id)
                print(f"  updated user '{email}' ({existing_id})")
        else:
            if dry_run:
                print(f"  [dry-run] create user '{email}'")
            else:
                new_id = _save_user(session, row, bu_id, "0")
                state.setdefault("users", {})[key] = new_id
                print(f"  created user '{email}' ({new_id})")


def _save_user(session, row, business_unit_id, edit_id):
    url = session.base_url + USER_EDIT_PATH
    form_id = session.get_form_id(url + f"?edit={edit_id}&refresh_parent=1&ext_edit=1")
    payload = user_payload(row, business_unit_id, form_id, edit_id)
    text = session.post_form(url, payload)
    m = re.search(r"edit=(\d+)", text)
    if edit_id == "0":
        if not m:
            raise SyncError(f"user create did not return an id: {text[:300]}")
        return m.group(1)
    return edit_id


def _delete_user(session, user_id):
    url = session.base_url + USER_DELETE_PATH
    form_id = session.get_form_id(url)
    session.post_form(
        url,
        {
            "checkbox[6]": user_id,
            "second_confirm": "yes",
            "confirmed": "yes",
            "owner": "",
            "_form_id": form_id,
            "dl[name]": "",
            "dl[title]": "",
            "dl[bu]": "-1",
            "dl[external_id]": "",
            "dl[uname]": "",
            "dl[last_login_date]": "",
            "dl[date]": "",
            "dl[offset]": "0",
            "dl[no_update]": "0",
            "dl[hide_columns]": "country,email,assigned_quota,external_id",
            "dl[column]": "0",
            "dl[order]": "1",
            "dl[rows_shown]": "20",
            "dl[row_selected]": "",
            "dl[quick_filter]": "",
            "dl[export]": "",
            "dl[row_select_available]": user_id,
            "search_hid": "0",
            "search_on": "",
            "dl_action": "",
        },
    )


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", help="Path to the Qualys xlsx workbook")
    parser.add_argument("--state", default="qualys_gui_state.json", help="State file path")
    parser.add_argument("--dry-run", action="store_true", help="Print actions without applying")
    args = parser.parse_args()

    wb = load_workbook(args.workbook, data_only=True)

    state = {"business_units": {}, "users": {}}
    if Path(args.state).exists():
        with open(args.state, encoding="utf-8") as fh:
            state = json.load(fh)

    if args.dry_run:
        print("DRY RUN - no changes will be applied")
        sync_business_units(None, wb, state, True)
        sync_users(None, wb, state, True)
        return

    session = QualysGUISession(GUI_URL, USERNAME, PASSWORD)
    session.login()
    print(f"logged in to {GUI_URL}")

    # Business units first so users can reference them by title.
    sync_business_units(session, wb, state, False)
    sync_users(session, wb, state, False)

    with open(args.state, "w", encoding="utf-8") as fh:
        json.dump(state, fh, indent=2, ensure_ascii=False)
    print(f"wrote state to {args.state}")


if __name__ == "__main__":
    try:
        main()
    except SyncError as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)
